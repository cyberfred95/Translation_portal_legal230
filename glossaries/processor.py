import csv
import io
import os.path
import openpyxl
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db.models.fields.files import FieldFile
import pandas as pd
from io import StringIO
from charset_normalizer import from_bytes
from django.core.exceptions import ValidationError


class GlossaryProcessor:

    @staticmethod
    def __get_csv_file_encoding(glossary_file):

        result = from_bytes(glossary_file.read())
        glossary_file.seek(0)
        return result[0].encoding

    @staticmethod
    def __check_on_duplicate(source_values: list, value, row_number):
        if value:
            if value not in source_values:
                source_values.append(value)
                return source_values
            else:
                raise ValidationError(f"Source value {value} is duplicated in line {row_number}")

    def convert_file_to_utf_8(self, csv_glossary_file):
        encoding = self.__get_csv_file_encoding(csv_glossary_file)

        file_content = csv_glossary_file.read().decode(encoding)
        csv_glossary_file.seek(0)
        try:
            file_stream = io.StringIO(file_content)
            df = pd.read_csv(file_stream)
        except pd.errors.ParserError as e:
            raise ValidationError(str(e))

        output_stream = io.StringIO()
        df.to_csv(output_stream, encoding="utf-8", index=False)
        output_stream.seek(0)

        converted_file = InMemoryUploadedFile(
            file=io.BytesIO(output_stream.getvalue().encode("utf-8")),
            field_name=getattr(csv_glossary_file, 'field_name', 'file'),
            name=csv_glossary_file.name,
            content_type="text/csv",
            size=len(output_stream.getvalue()),
            charset="utf-8",
        )
        return converted_file

    @staticmethod
    def __check_on_unsupported_symbols(row: list, row_number: int):
        for column in row[:1]:
            if column:
                if column.startswith("#"):
                    try:
                        column.encode('utf-8').decode('utf-8')
                    except UnicodeDecodeError as e:
                        raise ValidationError(
                            f"Invalid UTF-8 character at position {e.start}: {column[e.start:e.end]} on line {row_number}"
                        )

    @staticmethod
    def __validate_on_empy_columns(row: list, row_number: int):
        if row[0] is None or row[0] == '' or row[0] == ' ' and (row[1] != '' or row[1] != ' '):
            raise ValidationError(
                f"Source column is blank at line {row_number}.")
        elif row[1] is None or row[1] == '' or row[1] == ' ' and (row[0] != '' or row[0] != ' '):
            raise ValidationError(
                f"Target column is blank at line {row_number}."
            )

    def __validate_csv_file(self, glossary_file):
        if isinstance(glossary_file, FieldFile):
            glossary_file.open('rb')

        try:
            # Automatically detect file encoding
            encoding = self.__get_csv_file_encoding(glossary_file)
            text_file = io.TextIOWrapper(glossary_file, encoding=encoding)
            source_values = []
            csv_reader = csv.reader(text_file)
            next(csv_reader, None)

            for row_number, row in enumerate(csv_reader, start=2):
                if len(row) < 2 or (len(row) == 3 and row[2] != '') or len(row) > 3:
                    raise ValidationError(f"Invalid row at line {row_number}: {row}. Expected two columns."
                                          )
                self.__validate_on_empy_columns(row=row, row_number=row_number)
                for column in row:
                    if column:
                        column = column.strip()
                self.__check_on_duplicate(source_values=source_values, value=row[0], row_number=row_number)
                self.__check_on_unsupported_symbols(row, row_number=row_number)

            text_file.detach()

        finally:
            if isinstance(glossary_file, FieldFile):
                glossary_file.close()

    def __validate_xlsx_file(self, glossary_file):
        if hasattr(glossary_file, 'file'):
            file_obj = glossary_file.file
        else:
            file_obj = glossary_file

        if hasattr(file_obj, 'seek'):
            try:
                file_obj.seek(0)
            except ValueError:
                glossary_file.open('rb')
                file_obj = glossary_file.file
                file_obj.seek(0)

        workbook = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = workbook.active
        source_values = []

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 2 or (len(row) == 3 and row[2]) or len(row) > 3:
                raise ValidationError(
                    f"Invalid row at line {row_number}: {row}. Expected two columns."
                )
            self.__validate_on_empy_columns(row=row, row_number=row_number)
            for column in row:
                if column:
                    column = column.strip()
            self.__check_on_duplicate(source_values=source_values, value=row[0], row_number=row_number)
            self.__check_on_unsupported_symbols(row, row_number=row_number)

    def validate_file(self, glossary_file):
        file_extension = os.path.splitext(glossary_file.name)[1]
        if file_extension == '.csv':
            self.__validate_csv_file(glossary_file)
        elif file_extension in ['.xlsx']:
            self.__validate_xlsx_file(glossary_file)
        else:
            raise ValidationError("Invalid file type")

    def __form_glossary_from_csv(self, glossary_file) -> list:
        value = []
        if isinstance(glossary_file, FieldFile):
            glossary_file.open('rb')
        
        try:
            # Automatically detect file encoding
            encoding = self.__get_csv_file_encoding(glossary_file)
            text_file = io.TextIOWrapper(glossary_file, encoding=encoding)
            csv_reader = csv.reader(text_file)
            next(csv_reader, None)

            for row in csv_reader:
                value.append([row[0], row[1]])

            text_file.detach()
            return value
            
        finally:
            if isinstance(glossary_file, FieldFile):
                glossary_file.close()

    @staticmethod
    def __form_glossary_from_xlsx(glossary_file) -> list:
        value = []
        with glossary_file.open(mode='rb') as file:
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and row[1] is not None:
                    value.append([row[0], row[1]])

        return value

    def form_glossary_object(self, glossary_file):
        self.validate_file(glossary_file)
        file_extension = os.path.splitext(glossary_file.name)[1]
        if file_extension == '.csv':
            return self.__form_glossary_from_csv(glossary_file=glossary_file)

        elif file_extension in ['.xlsx', '.xls']:
            return self.__form_glossary_from_xlsx(glossary_file=glossary_file)
